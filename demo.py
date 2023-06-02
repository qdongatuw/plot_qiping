import openpyxl
from openpyxl.utils import get_column_letter

# Load the Excel file
workbook = openpyxl.load_workbook('output_mecp2_0531.xlsx')


# Select worksheet
worksheets = workbook.sheetnames

for i in worksheets:
    worksheet = workbook[i]
    # Loop through each row and column
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    for column in range(1, worksheet.max_column + 1):

        for row in reversed(list(range(1, worksheet.max_row + 1))):
                    # Check if cell is empty
            if not worksheet.cell(row=row, column=column).value:
                # Delete the cell
                # worksheet.delete_rows(row, amount=1)
                # Shift cells up
                print(f'{get_column_letter(column)}{row+1}:{get_column_letter(column)}{max_row}')
                worksheet.move_range(f'{get_column_letter(column)}{row+1}:{get_column_letter(column)}{max_row+1}', rows=-1)
                

# Save the changes
workbook.save('modified.xlsx')
